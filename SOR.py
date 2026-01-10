import streamlit as st
import numpy as np
import matplotlib.pyplot as plt
import sqlite3
import bcrypt
from datetime import datetime as dt_datetime, date, time
import time as time_module
from PIL import Image
import os
import base64
import pandas as pd
from collections import defaultdict
import io
import zipfile
import tempfile
import shutil
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.drawing.image import Image as ExcelImage

from Utils import *
from Css import applay_Css

def show_sor_page():
    """Page SOR (Safety Observation Report) pour HSE et Chef de chantier"""
    applay_Css()
    
    if 'sor_list' not in st.session_state:
        st.session_state.sor_list = []
    
    if 'sor_counter' not in st.session_state:
        st.session_state.sor_counter = 1
    
    if 'sor_photos' not in st.session_state:
        st.session_state.sor_photos = {}
    
    if 'annexes_list' not in st.session_state:
        st.session_state.annexes_list = []
    
    # Nouvel état pour stocker les SOR avec photos du tab1
    if 'sor_with_photos_buffer' not in st.session_state:
        st.session_state.sor_with_photos_buffer = []
        

    display_app_header("Safety Observation Report (SOR)")
    
    user = st.session_state.current_user
    user_name = get_user_name(user)
    user_statut = get_user_status(user)
    
    emplacements_bruts = [
        "Hall 1 & Hall 2 Exterior works",
        "H1-Fondations",
        "H1-Caissons & Purlins",
        "H1-Arches",
        "H2-Arches",
        "H1- Exterior works",
        "H1-Exterior works",
        "H1-Interior GC Works",
        "H2-Fondations",
        "H2-Caissons & Purlins",
        "H2-Exterior works",
        "H2-Interior GC Works",
        "Roads & Networks"
    ]
    
    types_intervention = ["SST", "Audit HSE", "SSTEPI", "Formation sécurité", 
                         "Réunion sécurité", "Inspection équipements", 
                         "Contrôle EPI", "Analyse risques", "Investigation accident",
                         "Exercice évacuation", "Visite de chantier", "Autre"]
    
    st.markdown("""
    <style>
        /* Supprime toutes les lignes horizontales */
        hr, .st-emotion-cache-1dp5vir {
            display: none !important;
        }
        
        /* Réduit l'espace après le titre */
        .st-emotion-cache-10trblm {
            margin-bottom: 0.5rem !important;
        }
        
        /* Réduit l'espace avant "SOR" */
        h2 {
            margin-top: 0.5rem !important;
        }
    </style>
    """, unsafe_allow_html=True)

    #styled_subheader("SOR")
    
    
    
    tab1, tab2, tab3 = st.tabs([
        "SOR JESA", 
        "Annexes avec Photos",
        "Statistiques"
    ])

    with tab1:
        st.markdown("### Préparer les SOR")
        
        with st.form(key="new_sor_form", clear_on_submit=False):
            col1, col2 = st.columns(2)
            with col1:
                sor_date = st.date_input("Date du SOR", value=date.today())
                chantier = st.selectbox("Emplacement :", emplacements_bruts)
                type_intervention = st.selectbox("Type d'intervention :", types_intervention)
            
            with col2:
                observateur = st.text_input("Observateur :", value=user_name, disabled=True)
                statut_sor = st.radio("Sévérité :", ["Risque Insignifiant", "Mineur", "Modéré", "Majeur", "Critique"], horizontal=True)
                priorite = st.select_slider("Priorité :", options=["Basse", "Moyenne", "Haute", "Critique"], value="Moyenne")

            col1, col2 = st.columns(2)
            with col1:
                description = st.text_area("Description détaillée *", help="Description obligatoire")
                st.markdown("##### PSP")
                c1, c2, c3 = st.columns(3)
                act1 = c1.checkbox("Formation", key="form_check")
                act2 = c2.checkbox("Conviction.Attitude", key="att_check")
                act3 = c3.checkbox("Ressources", key="res_check")
                
                actions = []
                if act1:
                    actions.append("Sensibilisation")
                if act2:
                    actions.append("Signalisation")
                if act3:
                    actions.append("Stop")
                actions_str = ", ".join(actions) if actions else "Aucune"
            
           
            with col2:
                st.markdown("####  Ajouter des photos au SOR")
                uploaded_files = st.file_uploader(
                    "Télécharger des photos (PNG, JPG, JPEG)",
                    type=["png", "jpg", "jpeg"],
                    accept_multiple_files=True,
                    key="sor_photos_uploader",
                    help="Vous pouvez sélectionner plusieurs photos"
                )
                
                if uploaded_files:
                    col_preview1, col_preview2 = st.columns(2)
                    with col_preview1:
                        st.markdown(f"**{len(uploaded_files)} photo(s) sélectionnée(s)**")
                    with col_preview2:
                     
                        preview_cols = st.columns(min(3, len(uploaded_files)))
                        for i, uploaded_file in enumerate(uploaded_files[:3]):
                            with preview_cols[i]:
                                st.image(uploaded_file, caption=f"Photo {i+1}", width=100)
                        if len(uploaded_files) > 3:
                            st.caption(f"... et {len(uploaded_files) - 3} autre(s) photo(s)")
            
            
            add_to_buffer = st.form_submit_button("➕ Ajouter à la liste de préparation", use_container_width=True)
                
            if add_to_buffer:
                    if not description:
                        st.error("La description est obligatoire")
                    else:
                        
                        sor_id = f"SOR-JESA-{dt_datetime.now().strftime('%Y%m%d')}-{st.session_state.sor_counter:03d}"
                        
                       
                        sor_entry = {
                            "ID": sor_id,
                            "Date": sor_date.strftime('%d/%m/%Y'),
                            "Observateur": user_name,
                            "Emplacement": chantier,
                            "Type Intervention": type_intervention,
                            "Sévérité": statut_sor,
                            "Priorité": priorite,
                            "Actions": actions_str,
                            "Description": description,
                            "Statut Traitement": "En cours",
                            "Date d'enregistrement": dt_datetime.now().strftime('%d/%m/%Y %H:%M'),
                            "Nombre de photos": len(uploaded_files) if uploaded_files else 0,
                            "Photos disponibles": "OUI" if uploaded_files else "NON"
                        }
                        
                      
                        st.session_state.sor_with_photos_buffer.append(sor_entry)
                        
                      
                        if uploaded_files:
                            photo_data = []
                            photo_filenames = []
                            for uploaded_file in uploaded_files:
                                bytes_data = uploaded_file.getvalue()
                                b64_encoded = base64.b64encode(bytes_data).decode()
                                
                                photo_data.append({
                                    "filename": uploaded_file.name,
                                    "data": b64_encoded,
                                    "type": uploaded_file.type,
                                    "size": len(bytes_data),
                                    "bytes": bytes_data
                                })
                                photo_filenames.append(uploaded_file.name)
                            
                            st.session_state.sor_photos[sor_id] = photo_data
                            sor_entry["Noms des photos"] = "; ".join(photo_filenames)
                        
                        st.session_state.sor_counter += 1
                        st.success(f"SOR {sor_id} ajouté à la liste de préparation !")
        
        # Afficher la liste des SOR en attente
        if st.session_state.sor_with_photos_buffer:
            
            st.markdown("###  Liste de préparation des SOR")
            
            # Fonction pour créer l'affichage des photos avec icônes
            def get_photo_display(photo_count, sor_id):
                if photo_count == 0:
                    return "📷 Aucune"
                elif photo_count == 1:
                    return "📸 1 photo"
                elif photo_count == 2:
                    return "📸📸 2 photos"
                elif photo_count == 3:
                    return "📸📸📸 3 photos"
                else:
                    return f"📸📸📸+ {photo_count} photos"
            
       
            buffer_data = []
            for sor in st.session_state.sor_with_photos_buffer:
                photo_count = sor["Nombre de photos"]
                photo_display = get_photo_display(photo_count, sor["ID"])
                
                buffer_data.append({
                    "ID": sor["ID"],
                    "Date": sor["Date"],
                    "Observateur": sor["Observateur"],
                    "Emplacement": sor["Emplacement"],
                    "Sévérité": sor["Sévérité"],
                    "Priorité": sor["Priorité"],
                    "Photos": photo_display,
                    "Description": sor["Description"][:50] + "..." if len(sor["Description"]) > 50 else sor["Description"]
                })
            
            buffer_df = pd.DataFrame(buffer_data)
            
            # Configuration des colonnes avec des icônes pour les photos
            st.dataframe(
                buffer_df,
                column_config={
                    "ID": st.column_config.TextColumn("ID", width=180),
                    "Date": st.column_config.TextColumn("Date", width=90),
                    "Observateur": st.column_config.TextColumn("Observateur", width=120),
                    "Emplacement": st.column_config.TextColumn("Emplacement", width=150),
                    "Sévérité": st.column_config.TextColumn("Sévérité", width=100),
                    "Priorité": st.column_config.TextColumn("Priorité", width=80),
                    "Photos": st.column_config.TextColumn("📸 Photos", width=120),
                    "Description": st.column_config.TextColumn("Description", width=200)
                },
                hide_index=True,
                use_container_width=True
            )
            
            # Boutons d'action
            col_save, col_export, col_clear = st.columns(3)
            
            with col_save:
                if st.button("Enregistrer TOUS les SOR", type="primary", use_container_width=True):
                    for sor in st.session_state.sor_with_photos_buffer:
                        # Ajouter aux listes principales
                        st.session_state.sor_list.append(sor)
                        st.session_state.annexes_list.append(sor)
                    
                    st.session_state.sor_with_photos_buffer = []
                    st.success(f"{len(buffer_data)} SOR enregistrés avec succès !")
                    time_module.sleep(1)
                    st.rerun()
            
            with col_export:
                # Créer et télécharger l'Excel avec photos
                excel_data = create_sor_excel_with_photos()
                if excel_data:
                    st.download_button(
                        label="Exporter vers Excel",
                        data=excel_data,
                        file_name=f"SOR_JESA_{dt_datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        help="Excel avec tous les SOR et leurs photos intégrées",
                        use_container_width=True
                    )
            
            with col_clear:
                if st.button("Vider la liste", use_container_width=True):
                    st.session_state.sor_with_photos_buffer = []
                    st.rerun()
            
            # Visualisation des détails d'un SOR spécifique
            st.markdown("---")
            st.markdown("#### Visualiser un SOR")
            
            sor_ids = [sor["ID"] for sor in st.session_state.sor_with_photos_buffer]
            selected_sor_id = st.selectbox(
                "Sélectionnez un SOR pour voir les détails:",
                options=[""] + sor_ids,
                key="select_sor_view"
            )
            
            if selected_sor_id:
                selected_sor = next((sor for sor in st.session_state.sor_with_photos_buffer if sor["ID"] == selected_sor_id), None)
                
                if selected_sor:
                    col_info, col_photos = st.columns([1, 1])
                    
                    with col_info:
                        st.markdown("**Informations du SOR:**")
                        st.markdown(f"**ID:** {selected_sor['ID']}")
                        st.markdown(f"**Observateur:** {selected_sor['Observateur']}")
                        st.markdown(f"**Date:** {selected_sor['Date']}")
                        st.markdown(f"**Emplacement:** {selected_sor['Emplacement']}")
                        st.markdown(f"**Type Intervention:** {selected_sor['Type Intervention']}")
                        st.markdown(f"**Sévérité:** {selected_sor['Sévérité']}")
                        st.markdown(f"**Priorité:** {selected_sor['Priorité']}")
                        st.markdown(f"**Actions PSP:** {selected_sor['Actions']}")
                        st.markdown(f"**Description:** {selected_sor['Description']}")
                    
                    with col_photos:
                        if selected_sor_id in st.session_state.sor_photos:
                            photo_list = st.session_state.sor_photos[selected_sor_id]
                            st.markdown(f"** 📸 Photos attachées ({len(photo_list)}):**")
                            
                            # Afficher les photos
                            cols = st.columns(min(3, len(photo_list)))
                            for idx, photo_info in enumerate(photo_list[:6]):
                                with cols[idx % 3]:
                                    img_data = base64.b64decode(photo_info["data"])
                                    image = Image.open(io.BytesIO(img_data))
                                    st.image(image, caption=f"{photo_info['filename']}", width=150)
                            
                            if len(photo_list) > 6:
                                st.caption(f"{len(photo_list) - 6} autre(s) photo(s) non affichée(s)")
                        else:
                            st.info("📷 Aucune photo attachée à ce SOR.")

    with tab2:
        st.markdown("###  Annexes - Enregistrement SOR avec Photos")
        
        with st.form(key="annexes_form", clear_on_submit=True):
            col1, col2 = st.columns(2)
            
            with col1:
                observateur_annexe = st.text_input(
                    "Observateur *", 
                    value=user_name,
                    help="Nom de la personne qui fait l'observation"
                )
                
                date_annexe = st.date_input(
                    "Date *",
                    value=date.today(),
                    help="Date de l'observation"
                )
            
            with col2:
                heure_annexe = st.time_input(
                    "Heure *",
                    value=dt_datetime.now().time(),
                    help="Heure de l'observation"
                )
                
                st.markdown("PSP *")
                psp_col1, psp_col2, psp_col3 = st.columns(3)
                with psp_col1:
                    psp_formation = st.checkbox("Formation")
                with psp_col2:
                    psp_attitude = st.checkbox("Conviction/Attitude")
                with psp_col3:
                    psp_ressources = st.checkbox("Ressources")
                
                psp_selection = []
                if psp_formation:
                    psp_selection.append("Formation")
                if psp_attitude:
                    psp_selection.append("Conviction/Attitude")
                if psp_ressources:
                    psp_selection.append("Ressources")
                
                psp_str = ", ".join(psp_selection) if psp_selection else "Aucun"
            
            st.markdown("---")
            st.markdown("#### Téléchargement de photos")
            
            col_photo1, col_photo2, col_photo3 = st.columns(3)
            
            with col_photo1:
                uploaded_files_annexe = st.file_uploader(
                    "Ajouter des photos (PNG, JPG, JPEG)",
                    type=["png", "jpg", "jpeg"],
                    accept_multiple_files=True,
                    help="Vous pouvez sélectionner plusieurs photos",
                    key="annexes_photos_uploader"
                )
            
            with col_photo2:
                if uploaded_files_annexe:
                    st.markdown(f"**{len(uploaded_files_annexe)} photo(s) sélectionnée(s)**")
                    for i, uploaded_file in enumerate(uploaded_files_annexe[:3]):
                        st.image(uploaded_file, caption=f"Photo {i+1}", width=100)
                    if len(uploaded_files_annexe) > 3:
                        st.caption(f"... et {len(uploaded_files_annexe) - 3} autre(s) photo(s)")
            with col_photo3:
                col_submit1, col_submit2 = st.columns([1, 2])
                with col_submit2:
                    submit_annexe = st.form_submit_button(
                        " Enregistrer l'annexe avec photos",
                        type="primary",
                        use_container_width=True
                    )
                
                if submit_annexe:
                    if not observateur_annexe:
                        st.error("Le champ 'Observateur' est obligatoire")
                    elif not psp_selection:
                        st.error("Veuillez sélectionner au moins un PSP")
                    else:
                        # Générer ID pour l'annexe
                        annexe_id = f"ANNEXE-{dt_datetime.now().strftime('%Y%m%d-%H%M%S')}"
                        
                        # Stocker les données
                        annexe_data = {
                            "ID": annexe_id,
                            "Observateur": observateur_annexe,
                            "Date": date_annexe.strftime('%d/%m/%Y'),
                            "Heure": heure_annexe.strftime('%H:%M'),
                            "PSP": psp_str,
                            "Date d'enregistrement": dt_datetime.now().strftime('%d/%m/%Y %H:%M:%S'),
                            "Nombre de photos": len(uploaded_files_annexe) if uploaded_files_annexe else 0,
                            "Photos disponibles": "OUI" if uploaded_files_annexe else "NON"
                        }
                        
                        # Stocker les photos
                        if uploaded_files_annexe:
                            photo_data = []
                            photo_filenames = []
                            for uploaded_file in uploaded_files_annexe:
                                bytes_data = uploaded_file.getvalue()
                                b64_encoded = base64.b64encode(bytes_data).decode()
                                
                                photo_data.append({
                                    "filename": uploaded_file.name,
                                    "data": b64_encoded,
                                    "type": uploaded_file.type,
                                    "size": len(bytes_data),
                                    "bytes": bytes_data
                                })
                                photo_filenames.append(uploaded_file.name)
                            
                            st.session_state.sor_photos[annexe_id] = photo_data
                            annexe_data["Noms des photos"] = "; ".join(photo_filenames)
                            annexe_data["Taille totale (KB)"] = f"{sum([p['size'] for p in photo_data]) / 1024:.1f}"
                        else:
                            annexe_data["Noms des photos"] = ""
                            annexe_data["Taille totale (KB)"] = "0"
                        
                        # Ajouter aux listes
                        st.session_state.annexes_list.append(annexe_data)
                        st.session_state.sor_list.append(annexe_data)
                        
                        st.success(f" 📸 Annexe {annexe_id} enregistrée avec succès !")
        
        st.markdown("###  Liste des annexes enregistrées")
        
        if not st.session_state.annexes_list:
            st.info("Aucune annexe n'a été enregistrée pour le moment.")
        else:
            # Fonction pour afficher les photos avec icônes
            def get_photo_display_annexe(photo_count):
                if photo_count == 0:
                    return "📷 Aucune"
                elif photo_count == 1:
                    return "📸 1"
                elif photo_count == 2:
                    return "📸📸 2"
                elif photo_count == 3:
                    return "📸📸📸 3"
                elif photo_count > 9:
                    return f"📸📸📸 {photo_count}"
                else:
                    # Créer une chaîne avec le bon nombre d'icônes
                    icons = "📸" * min(photo_count, 3)
                    if photo_count > 3:
                        icons += f" +{photo_count-3}"
                    return icons
            
            # Préparer les données avec affichage des icônes
            display_data = []
            for annexe in st.session_state.annexes_list:
                photo_count = annexe.get("Nombre de photos", 0)
                photo_display = get_photo_display_annexe(photo_count)
                
                display_data.append({
                    "ID": annexe.get("ID", ""),
                    "Date": annexe.get("Date", ""),
                    "Heure": annexe.get("Heure", ""),
                    "Observateur": annexe.get("Observateur", ""),
                    "PSP": annexe.get("PSP", ""),
                    "📸 Photos": photo_display,
                    "Enregistré le": annexe.get("Date d'enregistrement", "")
                })
            
            display_df = pd.DataFrame(display_data)
            
            st.dataframe(
                display_df,
                column_config={
                    "ID": st.column_config.TextColumn("ID", width=200),
                    "Date": st.column_config.TextColumn("Date", width=90),
                    "Heure": st.column_config.TextColumn("Heure", width=80),
                    "Observateur": st.column_config.TextColumn("Observateur", width=120),
                    "PSP": st.column_config.TextColumn("PSP", width=150),
                    "📸 Photos": st.column_config.TextColumn("📸 Photos", width=100),
                    "Enregistré le": st.column_config.TextColumn("Enregistré le", width=150)
                },
                hide_index=True,
                use_container_width=True,
                height=300
            )
            
            st.markdown("#### Exporter les SOR avec Photos")
            
            col_export1, col_export2 = st.columns(2)
            
            with col_export1:
                excel_data = create_excel_with_embedded_photos()
                if excel_data:
                    st.download_button(
                        label=" 📸 Excel avec Photos INTÉGRÉES",
                        data=excel_data,
                        file_name=f"SOR_avec_photos_integres_{dt_datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        help="Excel unique avec TOUTES les photos intégrées - Chaque SOR avec ses photos",
                        use_container_width=True,
                        type="primary"
                    )
            
            with col_export2:
                zip_data = create_complete_zip_export()
                if zip_data:
                    st.download_button(
                        label=" 📦 ZIP de sauvegarde",
                        data=zip_data,
                        file_name=f"SOR_backup_{dt_datetime.now().strftime('%Y%m%d_%H%M%S')}.zip",
                        mime="application/zip",
                        help="Archive ZIP avec fichiers séparés (pour backup)",
                        use_container_width=True
                    )
            
            st.markdown("####  Visualiser une annexe")
            selected_annexe_id = st.selectbox(
                "Sélectionnez une annexe pour voir les détails:",
                options=[""] + [annexe["ID"] for annexe in st.session_state.annexes_list],
                key="select_annexe_view"
            )
            
            if selected_annexe_id:
                selected_annexe = next((annexe for annexe in st.session_state.annexes_list if annexe["ID"] == selected_annexe_id), None)
                
                if selected_annexe:
                    col_info, col_photos = st.columns([1, 1])
                    
                    with col_info:
                        st.markdown("**Informations de l'annexe:**")
                        st.markdown(f"**ID:** {selected_annexe['ID']}")
                        st.markdown(f"**Observateur:** {selected_annexe['Observateur']}")
                        st.markdown(f"**Date:** {selected_annexe['Date']}")
                        st.markdown(f"**Heure:** {selected_annexe['Heure']}")
                        st.markdown(f"**PSP:** {selected_annexe['PSP']}")
                        st.markdown(f"**Enregistré le:** {selected_annexe.get('Date d\'enregistrement', 'N/A')}")
                        
                        photo_count = selected_annexe.get('Nombre de photos', 0)
                        # Afficher les icônes pour le nombre de photos
                        if photo_count == 0:
                            photo_display = "📷 Aucune photo"
                        elif photo_count == 1:
                            photo_display = "📸 1 photo"
                        else:
                            photo_display = f"📸📸📸 {photo_count} photos"
                        st.markdown(f"**Nombre de photos:** {photo_display}")
                        
                        if selected_annexe.get('Noms des photos'):
                            st.markdown("**Noms des photos:**")
                            for photo_name in selected_annexe.get('Noms des photos', '').split('; '):
                                st.markdown(f"- 📷 {photo_name}")
                    
                    with col_photos:
                        if selected_annexe_id in st.session_state.sor_photos:
                            photo_list = st.session_state.sor_photos[selected_annexe_id]
                            st.markdown(f"** 📸 Photos attachées ({len(photo_list)}):**")
                            
                            cols = st.columns(min(3, len(photo_list)))
                            for idx, photo_info in enumerate(photo_list[:6]):
                                with cols[idx % 3]:
                                    img_data = base64.b64decode(photo_info["data"])
                                    image = Image.open(io.BytesIO(img_data))
                                    st.image(image, caption=f"{photo_info['filename']} ({photo_info['size']/1024:.1f} KB)", width=150)
                            
                            if len(photo_list) > 6:
                                st.caption(f"{len(photo_list) - 6} autre(s) photo(s) non affichée(s)")
                        else:
                            st.info("📷 Aucune photo attachée à cette annexe.")

    with tab3:
        st.markdown("### Statistiques SOR")
        
        # Combiner toutes les listes pour les statistiques
        all_records = st.session_state.sor_list + st.session_state.annexes_list
        
        if not all_records:
            st.info("Aucune statistique disponible - Aucun SOR/Annexe enregistré")
        else:
            sor_df = pd.DataFrame(all_records)
            
            col_kpi1, col_kpi2, col_kpi3, col_kpi4 = st.columns(4)
            
            with col_kpi1:
                total_sor = len(all_records)
                st.metric("📋 Total SOR/Annexes", total_sor)
            
            with col_kpi2:
                if 'Priorité' in sor_df.columns:
                    sor_critique = len(sor_df[sor_df.get('Priorité', '') == 'Critique'])
                else:
                    sor_critique = 0
                st.metric("⚠️ Priorité Critique", sor_critique)
            
            with col_kpi3:
                if 'Emplacement' in sor_df.columns:
                    emplacements_count = len(sor_df['Emplacement'].unique())
                else:
                    emplacements_count = 0
                st.metric("📍 Emplacements", emplacements_count)
            
            with col_kpi4:
                total_photos = sum([record.get('Nombre de photos', 0) for record in all_records])
                st.metric("📸 Photos", total_photos)
            
            col_chart1, col_chart2 = st.columns(2)
            
            with col_chart1:
                st.markdown("##### Répartition par date")
                if 'Date' in sor_df.columns and not sor_df.empty:
                    try:
                        sor_df['Date_dt'] = pd.to_datetime(sor_df['Date'], format='%d/%m/%Y', errors='coerce')
                        date_counts = sor_df['Date_dt'].dt.date.value_counts().sort_index()
                        
                        if not date_counts.empty:
                            fig1, ax1 = plt.subplots(figsize=(10, 6))
                            ax1.plot(date_counts.index.astype(str), date_counts.values, marker='o', color='#004890')
                            ax1.set_xlabel('Date')
                            ax1.set_ylabel('Nombre de SOR/Annexes')
                            ax1.set_title('SOR/Annexes par date')
                            plt.xticks(rotation=45, ha='right')
                            st.pyplot(fig1)
                    except:
                        st.info("Données de date non disponibles pour le graphique")
            
            with col_chart2:
                st.markdown("##### Répartition par observateur")
                if 'Observateur' in sor_df.columns and not sor_df.empty:
                    observateur_counts = sor_df['Observateur'].value_counts()
                    if not observateur_counts.empty:
                        fig2, ax2 = plt.subplots(figsize=(10, 6))
                        bars = ax2.bar(observateur_counts.index, observateur_counts.values, color='#004890')
                        ax2.set_xlabel('Observateur')
                        ax2.set_ylabel('Nombre')
                        ax2.set_title('SOR/Annexes par observateur')
                        plt.xticks(rotation=45, ha='right')
                        st.pyplot(fig2)


# Les fonctions create_sor_excel_with_photos(), create_excel_with_embedded_photos(), 
# create_complete_zip_export() et send_email_with_excel() restent identiques
# (je ne les ai pas copiées ici pour garder le code concis)



def create_sor_excel_with_photos():
    """Créer un fichier Excel pour les SOR du tab1 avec photos intégrées"""
    if not st.session_state.sor_with_photos_buffer:
        return None
    
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "SOR JESA"
        
        # Styles
        header_fill = PatternFill(start_color="004890", end_color="004890", fill_type="solid")
        header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        
        # En-têtes
        headers = [
            "ID SOR", "Date", "Observateur", "Emplacement", "Type Intervention",
            "Sévérité", "Priorité", "Actions PSP", "Description", "Photos intégrées"
        ]
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
            
            # Ajuster la largeur des colonnes
            if header == "ID SOR":
                ws.column_dimensions[get_column_letter(col_idx)].width = 20
            elif header == "Description":
                ws.column_dimensions[get_column_letter(col_idx)].width = 40
            elif header == "Emplacement":
                ws.column_dimensions[get_column_letter(col_idx)].width = 25
            elif header == "Actions PSP":
                ws.column_dimensions[get_column_letter(col_idx)].width = 15
            elif header == "Photos intégrées":
                ws.column_dimensions[get_column_letter(col_idx)].width = 15
            else:
                ws.column_dimensions[get_column_letter(col_idx)].width = 12
        
        # Remplir les données
        for row_idx, sor in enumerate(st.session_state.sor_with_photos_buffer, start=2):
            sor_id = sor.get("ID", "")
            
            # Données de base
            ws.cell(row=row_idx, column=1, value=sor_id)
            ws.cell(row=row_idx, column=2, value=sor.get("Date", ""))
            ws.cell(row=row_idx, column=3, value=sor.get("Observateur", ""))
            ws.cell(row=row_idx, column=4, value=sor.get("Emplacement", ""))
            ws.cell(row=row_idx, column=5, value=sor.get("Type Intervention", ""))
            ws.cell(row=row_idx, column=6, value=sor.get("Sévérité", ""))
            ws.cell(row=row_idx, column=7, value=sor.get("Priorité", ""))
            ws.cell(row=row_idx, column=8, value=sor.get("Actions", ""))
            ws.cell(row=row_idx, column=9, value=sor.get("Description", ""))
            
            # Ajuster la hauteur de la ligne pour les photos
            ws.row_dimensions[row_idx].height = 100
            
            # Intégrer les photos
            if sor_id in st.session_state.sor_photos:
                photo_list = st.session_state.sor_photos[sor_id]
                
                if photo_list:
                    # Colonne pour les photos
                    photo_col = 10
                    
                    # Intégrer jusqu'à 2 photos par SOR
                    for idx, photo_info in enumerate(photo_list[:2]):
                        try:
                            # Décoder l'image
                            img_data = base64.b64decode(photo_info["data"])
                            img_pil = Image.open(io.BytesIO(img_data))
                            
                            # Redimensionner
                            max_size = (120, 120)
                            img_pil.thumbnail(max_size, Image.Resampling.LANCZOS)
                            
                            # Convertir pour Excel
                            img_buffer = io.BytesIO()
                            if img_pil.mode != 'RGB':
                                img_pil = img_pil.convert('RGB')
                            img_pil.save(img_buffer, format='JPEG', quality=85)
                            img_buffer.seek(0)
                            
                            # Créer l'image Excel
                            img_excel = ExcelImage(img_buffer)
                            img_excel.width = 100
                            img_excel.height = 80
                            
                            # Placer l'image
                            target_col = photo_col + idx
                            cell_ref = f"{get_column_letter(target_col)}{row_idx}"
                            ws.add_image(img_excel, cell_ref)
                            
                            # Ajuster la largeur de la colonne
                            ws.column_dimensions[get_column_letter(target_col)].width = 15
                            
                        except Exception as photo_error:
                            continue
                    
                    # Note si plus de 2 photos
                    if len(photo_list) > 2:
                        note_cell = ws.cell(row=row_idx, column=photo_col + 2, 
                                           value=f"+{len(photo_list) - 2} autre(s)")
                        note_cell.font = Font(size=8, italic=True)
                        note_cell.alignment = Alignment(vertical="center")
                
                else:
                    ws.cell(row=row_idx, column=10, value="Aucune")
            else:
                ws.cell(row=row_idx, column=10, value="Aucune")
            
            # Appliquer les bordures aux cellules de données
            for col_idx in range(1, 10):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = border
                cell.alignment = Alignment(vertical="center", wrap_text=True)
        
        # Créer une feuille de résumé
        ws_summary = wb.create_sheet(title="Résumé")
        
        ws_summary.merge_cells('A1:D1')
        title_cell = ws_summary.cell(row=1, column=1, value="RÉSUMÉ SOR JESA")
        title_cell.font = Font(name="Arial", size=14, bold=True, color="004890")
        title_cell.alignment = Alignment(horizontal="center")
        
        # Statistiques
        total_photos = sum([sor.get('Nombre de photos', 0) for sor in st.session_state.sor_with_photos_buffer])
        
        summary_data = [
            ["Date d'export", dt_datetime.now().strftime('%d/%m/%Y %H:%M:%S')],
            ["Nombre de SOR", len(st.session_state.sor_with_photos_buffer)],
            ["Total photos", total_photos],
            ["Photos intégrées", f"{min(total_photos, len(st.session_state.sor_with_photos_buffer) * 2)} (max 2 par SOR)"],
            ["Premier SOR", st.session_state.sor_with_photos_buffer[0]['ID'] if st.session_state.sor_with_photos_buffer else "N/A"],
            ["Dernier SOR", st.session_state.sor_with_photos_buffer[-1]['ID'] if st.session_state.sor_with_photos_buffer else "N/A"],
            ["", ""],
            ["NOTE:", "Les photos sont intégrées dans le fichier Excel."],
            ["", "Ouvrir avec Microsoft Excel pour voir les photos."]
        ]
        
        for row_idx, row_data in enumerate(summary_data, start=3):
            for col_idx, cell_value in enumerate(row_data, start=1):
                cell = ws_summary.cell(row=row_idx, column=col_idx, value=cell_value)
                cell.border = border
        
        ws_summary.column_dimensions['A'].width = 30
        ws_summary.column_dimensions['B'].width = 40
        
        # Sauvegarder dans le buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer.getvalue()
        
    except Exception as e:
        st.error(f"Erreur lors de la création de l'Excel SOR JESA: {str(e)}")
        return None


def create_excel_with_embedded_photos():
    """Créer un fichier Excel avec TOUTES les photos intégrées - Version corrigée"""
    if not st.session_state.annexes_list:
        return None
    
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "SOR avec Photos"
        
        header_fill = PatternFill(start_color="004890", end_color="004890", fill_type="solid")
        header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        
        headers = [
            "ID SOR", "Observateur", "Date", "Heure", "PSP", 
            "Date enregistrement", "Nb photos", "Photos intégrées"
        ]
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
            
            if header == "ID SOR":
                ws.column_dimensions[get_column_letter(col_idx)].width = 25
            elif header == "Observateur":
                ws.column_dimensions[get_column_letter(col_idx)].width = 20
            elif header == "PSP":
                ws.column_dimensions[get_column_letter(col_idx)].width = 20
            elif header == "Date enregistrement":
                ws.column_dimensions[get_column_letter(col_idx)].width = 18
            elif header == "Photos intégrées":
                ws.column_dimensions[get_column_letter(col_idx)].width = 15
            else:
                ws.column_dimensions[get_column_letter(col_idx)].width = 12
        
        for row_idx, annexe in enumerate(st.session_state.annexes_list, start=2):
            annexe_id = annexe.get("ID", "")
            
            ws.cell(row=row_idx, column=1, value=annexe_id)  
            ws.cell(row=row_idx, column=2, value=annexe.get("Observateur", ""))  
            ws.cell(row=row_idx, column=3, value=annexe.get("Date", ""))  
            ws.cell(row=row_idx, column=4, value=annexe.get("Heure", ""))  
            ws.cell(row=row_idx, column=5, value=annexe.get("PSP", ""))  
            ws.cell(row=row_idx, column=6, value=annexe.get("Date d'enregistrement", "")) 
            ws.cell(row=row_idx, column=7, value=annexe.get("Nombre de photos", 0))  
            
            ws.row_dimensions[row_idx].height = 80
            
            if annexe_id in st.session_state.sor_photos:
                photo_list = st.session_state.sor_photos[annexe_id]
                
                if photo_list:
                    photo_col = 8
                    
                    for idx, photo_info in enumerate(photo_list[:2]):
                        try:
                            img_data = base64.b64decode(photo_info["data"])
                            img_pil = Image.open(io.BytesIO(img_data))
                            
                            max_size = (100, 100)
                            img_pil.thumbnail(max_size, Image.Resampling.LANCZOS)
                            
                            img_buffer = io.BytesIO()
                            
                            if img_pil.mode != 'RGB':
                                img_pil = img_pil.convert('RGB')
                            img_pil.save(img_buffer, format='JPEG', quality=85)
                            img_buffer.seek(0)
                            
                            img_excel = ExcelImage(img_buffer)
                            img_excel.width = 80  
                            img_excel.height = 60  
                            
                            target_col = photo_col + idx
                            cell_ref = f"{get_column_letter(target_col)}{row_idx}"
                            ws.add_image(img_excel, cell_ref)
                            
                            ws.column_dimensions[get_column_letter(target_col)].width = 15
                            
                            if row_idx + 1 <= ws.max_row:
                                ws.row_dimensions[row_idx + 1].height = 20
                            caption_cell = ws.cell(row=row_idx + 1, column=target_col, value=f"P{idx+1}")
                            caption_cell.alignment = Alignment(horizontal="center")
                            caption_cell.font = Font(size=8)
                            
                        except Exception as photo_error:
                            error_cell = ws.cell(row=row_idx, column=photo_col + idx, value=f"Photo {idx+1}")
                            error_cell.font = Font(size=8, color="FF0000")
                            continue
                    
                    if len(photo_list) > 2:
                        remaining = len(photo_list) - 2
                        note_cell = ws.cell(row=row_idx, column=photo_col + 2, 
                                           value=f"+{remaining} autre(s)")
                        note_cell.font = Font(size=8, italic=True)
                        note_cell.alignment = Alignment(vertical="center")
                
                else:
                    ws.cell(row=row_idx, column=8, value="Aucune photo")
            else:
                ws.cell(row=row_idx, column=8, value="Aucune photo")
            
            for col_idx in range(1, 8): 
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = border
                cell.alignment = Alignment(vertical="center")
        
        ws_list = wb.create_sheet(title="Liste Photos")
        
        headers_list = ["ID SOR", "Observateur", "Date", "Heure", "Nom photo", "Taille (KB)", "Statut"]
        ws_list.append(headers_list)
        
        for col in range(1, len(headers_list) + 1):
            cell = ws_list.cell(row=1, column=col)
            cell.fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")
            cell.font = Font(bold=True)
            cell.border = border
        
        row_num = 2
        for annexe in st.session_state.annexes_list:
            annexe_id = annexe.get("ID", "")
            if annexe_id in st.session_state.sor_photos:
                photo_list = st.session_state.sor_photos[annexe_id]
                for idx, photo_info in enumerate(photo_list):
                    status = "Intégrée" if idx < 2 else "Non intégrée (limite 2/SOR)"
                    
                    ws_list.append([
                        annexe_id,
                        annexe.get("Observateur", ""),
                        annexe.get("Date", ""),
                        annexe.get("Heure", ""),
                        photo_info['filename'],
                        f"{photo_info['size']/1024:.1f}",
                        status
                    ])
                    
                    for col in range(1, len(headers_list) + 1):
                        cell = ws_list.cell(row=row_num, column=col)
                        cell.border = border
                    
                    row_num += 1
        
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            ws_list.column_dimensions[col].width = 20
        
        ws_summary = wb.create_sheet(title="Résumé")
        
        ws_summary.merge_cells('A1:C1')
        title_cell = ws_summary.cell(row=1, column=1, value="RÉSUMÉ SOR AVEC PHOTOS")
        title_cell.font = Font(name="Arial", size=14, bold=True, color="004890")
        title_cell.alignment = Alignment(horizontal="center")
        
        total_photos = sum([annexe.get('Nombre de photos', 0) for annexe in st.session_state.annexes_list])
        
        summary_data = [
            ["Date d'export", dt_datetime.now().strftime('%d/%m/%Y %H:%M:%S')],
            ["Nombre de SOR", len(st.session_state.annexes_list)],
            ["Total photos", total_photos],
            ["Photos intégrées", f"{min(total_photos, len(st.session_state.annexes_list) * 2)} (max 2 par SOR)"],
            ["Premier SOR", st.session_state.annexes_list[0]['ID'] if st.session_state.annexes_list else "N/A"],
            ["Dernier SOR", st.session_state.annexes_list[-1]['ID'] if st.session_state.annexes_list else "N/A"],
            ["", ""],
            ["FEUILLES:", ""],
            ["1. 'SOR avec Photos'", "Liste principale avec photos intégrées"],
            ["2. 'Liste Photos'", "Détail de toutes les photos"],
            ["3. 'Résumé'", "Cette page"],
            ["", ""],
            ["NOTE:", "Les photos sont intégrées dans le fichier Excel."],
            ["", "Ouvrir avec Microsoft Excel pour voir les photos."]
        ]
        
        for row_idx, row_data in enumerate(summary_data, start=3):
            for col_idx, cell_value in enumerate(row_data, start=1):
                cell = ws_summary.cell(row=row_idx, column=col_idx, value=cell_value)
                cell.border = border
                if row_idx >= 10:  
                    cell.font = Font(bold=(col_idx == 1))
        
        ws_summary.column_dimensions['A'].width = 30
        ws_summary.column_dimensions['B'].width = 40
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer.getvalue()
        
    except Exception as e:
        st.error(f"Erreur lors de la création de l'Excel: {str(e)}")
        return None


def create_complete_zip_export():
    """Créer un fichier ZIP complet pour backup"""
    if not st.session_state.annexes_list:
        return None
    
    zip_buffer = io.BytesIO()
    
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        excel_with_photos = create_excel_with_embedded_photos()
        if excel_with_photos:
            zip_file.writestr("SOR_avec_Photos.xlsx", excel_with_photos)
        
        photo_count = 0
        for annexe in st.session_state.annexes_list:
            annexe_id = annexe.get("ID", "")
            if annexe_id in st.session_state.sor_photos:
                photo_list = st.session_state.sor_photos[annexe_id]
                
                for idx, photo_info in enumerate(photo_list):
                    img_bytes = photo_info.get("bytes", base64.b64decode(photo_info["data"]))
                    filename = f"Photos_Originales/{annexe_id}/{photo_info['filename']}"
                    zip_file.writestr(filename, img_bytes)
                    photo_count += 1
        
        readme_content = f"""ARCHIVE SOR COMPLÈTE
========================

Date: {dt_datetime.now().strftime('%d/%m/%Y %H:%M:%S')}
SOR: {len(st.session_state.annexes_list)}
Photos: {photo_count}

Fichier principal: SOR_avec_Photos.xlsx
- Contient toutes les photos intégrées
- Ouvrir avec Microsoft Excel"""
        
        zip_file.writestr("LISEZ_MOI.txt", readme_content)
    
    zip_buffer.seek(0)
    return zip_buffer.getvalue()


def send_email_with_excel(to_email, subject, excel_data, filename):
    """Envoyer un email avec fichier Excel en pièce jointe"""
    try:
        import smtplib
        from email.mime.multipart import MIMEMultipart
        from email.mime.base import MIMEBase
        from email.mime.text import MIMEText
        from email import encoders
        
        GMAIL_USER = os.environ.get("GMAIL_USER", "consular.services.infos@gmail.com")
        GMAIL_PASSWORD = os.environ.get("GMAIL_PASSWORD", "ocwo vozn pskp copd")
        
        msg = MIMEMultipart()
        msg['From'] = GMAIL_USER
        msg['To'] = to_email
        msg['Subject'] = subject
        
        message_body = f"""Bonjour,

Veuillez trouver ci-joint le fichier Excel contenant les SOR avec leurs photos intégrées.

Fichier: {filename}
SOR: {len(st.session_state.annexes_list)}
Photos: {sum([annexe.get('Nombre de photos', 0) for annexe in st.session_state.annexes_list])}
Date: {dt_datetime.now().strftime('%d/%m/%Y %H:%M')}

Les photos sont intégrées directement dans le fichier Excel.
Ouvrir avec Microsoft Excel pour visualiser.

Cordialement,
Système HSE"""
        
        msg.attach(MIMEText(message_body, 'plain'))
        
        attachment = MIMEBase('application', 'octet-stream')
        attachment.set_payload(excel_data)
        encoders.encode_base64(attachment)
        attachment.add_header('Content-Disposition', f'attachment; filename="{filename}"')
        msg.attach(attachment)
        
        with smtplib.SMTP('smtp.gmail.com', 587) as server:
            server.starttls()
            server.login(GMAIL_USER, GMAIL_PASSWORD)
            server.send_message(msg)
        
        return True
        
    except Exception as e:
        st.error(f" Erreur d'envoi d'email: {str(e)}")
        return False